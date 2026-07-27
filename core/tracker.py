"""Typed, validated, and atomically persisted Excel application tracking."""

from __future__ import annotations

import logging
import os
import re
import threading
import unicodedata
import uuid
import zipfile
from collections.abc import Callable
from dataclasses import dataclass, replace
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Literal, Protocol
from xml.etree.ElementTree import ParseError

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet

from config import Settings
from core.url_safety import validate_http_url

LOGGER = logging.getLogger(__name__)

APPLICATION_COLUMNS: tuple[str, ...] = (
    "application_id",
    "created_at",
    "updated_at",
    "applied_date",
    "company",
    "role",
    "reference_number",
    "location",
    "job_url",
    "language",
    "fit_assessment",
    "status",
    "contact_person",
    "letter_path",
    "source_hash",
    "input_hash",
    "cv_version_id",
    "cv_reference_hash",
    "used_previous_cv",
    "notes",
)

_APPLICATIONS_SHEET = "Applications"
_SHA256_PATTERN = re.compile(r"[0-9a-f]{64}\Z")
_MAX_APPLICATION_ID_CHARS = 128
_MAX_CELL_TEXT_CHARS = 32_767
_LOCKED_WINERRORS = frozenset((5, 32, 33))


class TrackerError(RuntimeError):
    """Base error for tracker validation, loading, or persistence."""


class TrackerValidationError(TrackerError):
    """Raised when a requested tracker change contains invalid values."""


class TrackerSchemaError(TrackerError):
    """Raised when an existing owned workbook or row is incompatible."""


class TrackerWorkbookLockedError(TrackerError):
    """Raised when Windows or Excel prevents safe workbook access."""


class TrackerApplicationNotFoundError(TrackerError):
    """Raised when a manual edit targets an unknown application."""


class GeneratedApplicationData(Protocol):
    """Generated metadata required to create or refresh a tracker row."""

    company: str
    role: str
    reference_number: str | None
    location: str | None
    job_url: str | None
    language: str
    fit_assessment: str
    contact_person: str | None
    source_hash: str
    input_hash: str
    cv_version_id: str
    cv_reference_hash: str
    used_previous_cv: bool


class _Lock(Protocol):
    def __enter__(self) -> object: ...

    def __exit__(
        self,
        exception_type: type[BaseException] | None,
        exception: BaseException | None,
        traceback: object | None,
    ) -> bool | None: ...


@dataclass(frozen=True, slots=True)
class ApplicationRecord:
    """One validated application row exposed to the local user interface."""

    application_id: str
    created_at: datetime
    updated_at: datetime
    applied_date: date | None
    company: str
    role: str
    reference_number: str | None
    location: str | None
    job_url: str | None
    language: Literal["de", "en"]
    fit_assessment: str
    status: str
    contact_person: str | None
    letter_path: str
    source_hash: str
    input_hash: str
    cv_version_id: str
    cv_reference_hash: str
    used_previous_cv: bool
    notes: str


@dataclass(frozen=True, slots=True)
class _GeneratedFields:
    company: str
    role: str
    reference_number: str | None
    location: str | None
    job_url: str | None
    language: Literal["de", "en"]
    fit_assessment: str
    contact_person: str | None
    source_hash: str
    input_hash: str
    cv_version_id: str
    cv_reference_hash: str
    used_previous_cv: bool


class ApplicationTracker:
    """Own and safely update the local applications workbook."""

    _settings: Settings
    _clock: Callable[[], datetime]
    _lock: _Lock

    def __init__(
        self,
        settings: Settings,
        clock: Callable[[], datetime] | None = None,
        lock: _Lock | None = None,
    ) -> None:
        """Initialize the tracker with explicit configuration and dependencies."""

        self._settings = settings
        self._clock = clock or (lambda: datetime.now(UTC))
        self._lock = lock or threading.RLock()

    def list_applications(self) -> tuple[ApplicationRecord, ...]:
        """Return every validated application in workbook row order."""

        with self._lock:
            if not self._settings.applications_path.exists():
                return ()
            workbook = self._load_existing_workbook()
            try:
                rows = self._validated_rows(workbook)
                return tuple(record for _, record in rows)
            finally:
                workbook.close()

    def upsert_generated(
        self,
        application_id: str,
        output: GeneratedApplicationData,
        letter_path: Path,
    ) -> ApplicationRecord:
        """Create or refresh generated fields while preserving manual fields."""

        valid_id = _validate_application_id(application_id)
        generated = self._validate_generated(output)
        stored_letter_path = self._normalize_letter_path(letter_path)
        timestamp = self._current_time()

        with self._lock:
            workbook = self._load_or_create_workbook()
            try:
                rows = self._validated_rows(workbook)
                existing = next(
                    (
                        (row_number, record)
                        for row_number, record in rows
                        if record.application_id == valid_id
                    ),
                    None,
                )
                if existing is None:
                    record = ApplicationRecord(
                        application_id=valid_id,
                        created_at=timestamp,
                        updated_at=timestamp,
                        applied_date=None,
                        company=generated.company,
                        role=generated.role,
                        reference_number=generated.reference_number,
                        location=generated.location,
                        job_url=generated.job_url,
                        language=generated.language,
                        fit_assessment=generated.fit_assessment,
                        status="Draft",
                        contact_person=generated.contact_person,
                        letter_path=stored_letter_path,
                        source_hash=generated.source_hash,
                        input_hash=generated.input_hash,
                        cv_version_id=generated.cv_version_id,
                        cv_reference_hash=generated.cv_reference_hash,
                        used_previous_cv=generated.used_previous_cv,
                        notes="",
                    )
                    row_number = max(
                        workbook[_APPLICATIONS_SHEET].max_row + 1,
                        2,
                    )
                else:
                    row_number, prior = existing
                    record = replace(
                        prior,
                        updated_at=timestamp,
                        company=generated.company,
                        role=generated.role,
                        reference_number=generated.reference_number,
                        location=generated.location,
                        job_url=generated.job_url,
                        language=generated.language,
                        fit_assessment=generated.fit_assessment,
                        contact_person=generated.contact_person,
                        letter_path=stored_letter_path,
                        source_hash=generated.source_hash,
                        input_hash=generated.input_hash,
                        cv_version_id=generated.cv_version_id,
                        cv_reference_hash=generated.cv_reference_hash,
                        used_previous_cv=generated.used_previous_cv,
                    )
                self._write_record(
                    workbook[_APPLICATIONS_SHEET],
                    row_number,
                    record,
                )
                self._save_atomic(workbook)
            finally:
                workbook.close()
        LOGGER.info("Application tracker workbook updated")
        return record

    def update_manual_fields(
        self,
        application_id: str,
        *,
        applied_date: date | None,
        status: str,
        notes: str,
    ) -> ApplicationRecord:
        """Explicitly save the three user-owned fields for one existing row."""

        valid_id = _validate_application_id(application_id)
        valid_date = _validate_requested_date(applied_date)
        valid_status = self._validate_status(status)
        valid_notes = _validate_text(notes, "notes", required=False)
        timestamp = self._current_time()

        with self._lock:
            if not self._settings.applications_path.exists():
                raise TrackerApplicationNotFoundError(
                    f"Application {valid_id} does not exist in the tracker."
                )
            workbook = self._load_existing_workbook()
            try:
                rows = self._validated_rows(workbook)
                existing = next(
                    (
                        (row_number, record)
                        for row_number, record in rows
                        if record.application_id == valid_id
                    ),
                    None,
                )
                if existing is None:
                    raise TrackerApplicationNotFoundError(
                        f"Application {valid_id} does not exist in the tracker."
                    )
                row_number, prior = existing
                record = replace(
                    prior,
                    updated_at=timestamp,
                    applied_date=valid_date,
                    status=valid_status,
                    notes=valid_notes,
                )
                self._write_record(
                    workbook[_APPLICATIONS_SHEET],
                    row_number,
                    record,
                )
                self._save_atomic(workbook)
            finally:
                workbook.close()
        LOGGER.info("Application tracker manual fields updated")
        return record

    def resolve_letter_path(self, stored_path: str) -> Path:
        """Resolve and revalidate one project-relative Markdown archive path."""

        if not isinstance(stored_path, str) or not stored_path:
            raise TrackerValidationError(
                "The letter path must be a non-empty project-relative path."
            )
        return self._resolve_relative_letter_path(Path(stored_path))

    def _load_or_create_workbook(self) -> OpenpyxlWorkbook:
        if self._settings.applications_path.exists():
            return self._load_existing_workbook()
        workbook = Workbook(iso_dates=True)
        sheet = workbook.active
        sheet.title = _APPLICATIONS_SHEET
        self._write_header(sheet)
        return workbook

    def _load_existing_workbook(self) -> OpenpyxlWorkbook:
        try:
            with self._settings.applications_path.open("rb") as stream:
                workbook = load_workbook(
                    stream,
                    data_only=False,
                    read_only=False,
                )
            return workbook
        except PermissionError as error:
            raise _locked_error() from error
        except OSError as error:
            if _is_locked_windows_error(error):
                raise _locked_error() from error
            raise TrackerSchemaError(
                "The applications workbook is invalid or corrupted. Restore a "
                "valid copy; AutoCover did not overwrite it."
            ) from error
        except (
            InvalidFileException,
            KeyError,
            ParseError,
            ValueError,
            zipfile.BadZipFile,
        ) as error:
            raise TrackerSchemaError(
                "The applications workbook is invalid or corrupted. Restore a "
                "valid copy; AutoCover did not overwrite it."
            ) from error

    def _validated_rows(
        self,
        workbook: OpenpyxlWorkbook,
    ) -> list[tuple[int, ApplicationRecord]]:
        if _APPLICATIONS_SHEET not in workbook.sheetnames:
            raise TrackerSchemaError(
                "The Applications sheet must use AutoCover's exact 20-column schema."
            )
        sheet = workbook[_APPLICATIONS_SHEET]
        headers = tuple(
            sheet.cell(1, index).value
            for index in range(1, max(sheet.max_column, len(APPLICATION_COLUMNS)) + 1)
        )
        if (
            sheet.max_column != len(APPLICATION_COLUMNS)
            or headers != APPLICATION_COLUMNS
        ):
            raise TrackerSchemaError(
                "The Applications sheet must use AutoCover's exact 20-column schema."
            )

        rows: list[tuple[int, ApplicationRecord]] = []
        seen_ids: set[str] = set()
        for row_number in range(2, sheet.max_row + 1):
            cells = tuple(
                sheet.cell(row_number, column_number)
                for column_number in range(1, len(APPLICATION_COLUMNS) + 1)
            )
            for cell in cells:
                if cell.hyperlink is not None:
                    raise TrackerSchemaError(
                        "Applications data contains an unsupported hyperlink."
                    )
                if cell.data_type == "f":
                    raise TrackerSchemaError(
                        "Applications data contains an unsupported formula."
                    )
            if all(cell.value is None for cell in cells):
                continue
            try:
                record = self._record_from_values(
                    tuple(cell.value for cell in cells)
                )
            except TrackerValidationError as error:
                raise TrackerSchemaError(
                    "The Applications workbook contains invalid tracker data."
                ) from error
            if record.application_id in seen_ids:
                raise TrackerSchemaError(
                    "Duplicate application ID found in Applications; repair the "
                    "workbook before continuing."
                )
            seen_ids.add(record.application_id)
            rows.append((row_number, record))
        return rows

    def _record_from_values(
        self,
        values: tuple[object, ...],
    ) -> ApplicationRecord:
        application_id = _validate_application_id(values[0])
        created_at = _parse_utc_timestamp(values[1], "created timestamp")
        updated_at = _parse_utc_timestamp(values[2], "updated timestamp")
        applied = _validate_stored_date(values[3])
        company = _validate_text(values[4], "company", required=True)
        role = _validate_text(values[5], "role", required=True)
        reference_number = _validate_optional_text(values[6], "reference number")
        location = _validate_optional_text(values[7], "location")
        job_url = _validate_optional_url(values[8])
        language = _validate_language(values[9])
        fit_assessment = _validate_text(
            values[10],
            "fit assessment",
            required=True,
        )
        status = self._validate_status(values[11])
        contact_person = _validate_optional_text(values[12], "contact person")
        letter_path = _validate_text(values[13], "letter path", required=True)
        resolved_letter = self.resolve_letter_path(letter_path)
        canonical_letter_path = resolved_letter.relative_to(
            self._settings.project_root.resolve()
        ).as_posix()
        if canonical_letter_path != letter_path:
            raise TrackerValidationError(
                "The stored letter path is not canonical project-relative text."
            )
        source_hash = _validate_hash(values[14], "source hash")
        input_hash = _validate_hash(values[15], "input hash")
        cv_version_id = _validate_text(
            values[16],
            "CV version ID",
            required=True,
        )
        cv_reference_hash = _validate_hash(
            values[17],
            "CV reference hash",
        )
        if type(values[18]) is not bool:
            raise TrackerValidationError(
                "The previous-CV tracker value must be a boolean."
            )
        notes_value = "" if values[19] is None else values[19]
        notes = _validate_text(notes_value, "notes", required=False)
        return ApplicationRecord(
            application_id=application_id,
            created_at=created_at,
            updated_at=updated_at,
            applied_date=applied,
            company=company,
            role=role,
            reference_number=reference_number,
            location=location,
            job_url=job_url,
            language=language,
            fit_assessment=fit_assessment,
            status=status,
            contact_person=contact_person,
            letter_path=letter_path,
            source_hash=source_hash,
            input_hash=input_hash,
            cv_version_id=cv_version_id,
            cv_reference_hash=cv_reference_hash,
            used_previous_cv=values[18],
            notes=notes,
        )

    def _validate_generated(
        self,
        output: GeneratedApplicationData,
    ) -> _GeneratedFields:
        company = _validate_text(
            getattr(output, "company", None),
            "company",
            required=True,
        )
        role = _validate_text(
            getattr(output, "role", None),
            "role",
            required=True,
        )
        reference_number = _validate_optional_text(
            getattr(output, "reference_number", None),
            "reference number",
        )
        location = _validate_optional_text(
            getattr(output, "location", None),
            "location",
        )
        job_url = _validate_optional_url(getattr(output, "job_url", None))
        language = _validate_language(getattr(output, "language", None))
        fit_assessment = _validate_text(
            getattr(output, "fit_assessment", None),
            "fit assessment",
            required=True,
        )
        contact_person = _validate_optional_text(
            getattr(output, "contact_person", None),
            "contact person",
        )
        source_hash = _validate_hash(
            getattr(output, "source_hash", None),
            "source hash",
        )
        input_hash = _validate_hash(
            getattr(output, "input_hash", None),
            "input hash",
        )
        cv_version_id = _validate_text(
            getattr(output, "cv_version_id", None),
            "CV version ID",
            required=True,
        )
        cv_reference_hash = _validate_hash(
            getattr(output, "cv_reference_hash", None),
            "CV reference hash",
        )
        used_previous_cv = getattr(output, "used_previous_cv", None)
        if type(used_previous_cv) is not bool:
            raise TrackerValidationError(
                "The previous-CV provenance value must be a boolean."
            )
        return _GeneratedFields(
            company=company,
            role=role,
            reference_number=reference_number,
            location=location,
            job_url=job_url,
            language=language,
            fit_assessment=fit_assessment,
            contact_person=contact_person,
            source_hash=source_hash,
            input_hash=input_hash,
            cv_version_id=cv_version_id,
            cv_reference_hash=cv_reference_hash,
            used_previous_cv=used_previous_cv,
        )

    def _normalize_letter_path(self, letter_path: Path) -> str:
        if not isinstance(letter_path, Path):
            raise TrackerValidationError(
                "The letter path must be a project-relative Path."
            )
        if letter_path.is_absolute():
            raise TrackerValidationError(
                "The letter path must be relative to the project directory."
            )
        resolved = self._resolve_relative_letter_path(letter_path)
        return resolved.relative_to(
            self._settings.project_root.resolve()
        ).as_posix()

    def _resolve_relative_letter_path(self, relative_path: Path) -> Path:
        if relative_path.is_absolute():
            raise TrackerValidationError(
                "The letter path must be relative to the project directory."
            )
        project_root = self._settings.project_root.resolve()
        letters_root = self._settings.letters_dir.resolve()
        resolved = (project_root / relative_path).resolve()
        try:
            resolved.relative_to(project_root)
        except ValueError as error:
            raise TrackerValidationError(
                "The letter path must stay inside the project directory."
            ) from error
        try:
            resolved.relative_to(letters_root)
        except ValueError as error:
            raise TrackerValidationError(
                "The letter archive must stay inside the letters directory."
            ) from error
        if not resolved.exists() or not resolved.is_file():
            raise TrackerValidationError(
                "The archived letter does not exist."
            )
        if resolved.suffix.casefold() != ".md":
            raise TrackerValidationError(
                "The archived letter must be a Markdown file."
            )
        return resolved

    def _validate_status(self, value: object) -> str:
        status = _validate_text(value, "application status", required=True)
        if status not in self._settings.application_statuses:
            raise TrackerValidationError(
                "Choose a valid application status."
            )
        return status

    def _current_time(self) -> datetime:
        value = self._clock()
        if (
            not isinstance(value, datetime)
            or value.tzinfo is None
            or value.utcoffset() is None
        ):
            raise TrackerValidationError(
                "The tracker clock must return a timezone-aware datetime."
            )
        return value.astimezone(UTC)

    def _write_header(self, sheet: Worksheet) -> None:
        for column_number, name in enumerate(APPLICATION_COLUMNS, start=1):
            _write_text_cell(sheet.cell(1, column_number), name)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:T1"

    def _write_record(
        self,
        sheet: Worksheet,
        row_number: int,
        record: ApplicationRecord,
    ) -> None:
        values: tuple[object, ...] = (
            record.application_id,
            record.created_at.isoformat(),
            record.updated_at.isoformat(),
            record.applied_date,
            record.company,
            record.role,
            record.reference_number,
            record.location,
            record.job_url,
            record.language,
            record.fit_assessment,
            record.status,
            record.contact_person,
            record.letter_path,
            record.source_hash,
            record.input_hash,
            record.cv_version_id,
            record.cv_reference_hash,
            record.used_previous_cv,
            record.notes,
        )
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column_number)
            cell.hyperlink = None
            if column_number == 4:
                cell.value = value
                cell.number_format = "yyyy-mm-dd"
            elif column_number == 19:
                cell.value = value
            elif value is None:
                cell.value = None
            else:
                _write_text_cell(cell, value)

    def _save_atomic(self, workbook: OpenpyxlWorkbook) -> None:
        target = self._settings.applications_path
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
        except OSError as error:
            if _is_locked_windows_error(error):
                raise _locked_error() from error
            raise TrackerError(
                "Could not create the private tracker directory."
            ) from error
        temporary = target.parent / f".applications-{uuid.uuid4().hex}.xlsx"
        try:
            workbook.iso_dates = True
            try:
                workbook.save(temporary)
            except PermissionError as error:
                raise _locked_error() from error
            except OSError as error:
                if _is_locked_windows_error(error):
                    raise _locked_error() from error
                raise TrackerError(
                    "Could not save the temporary applications workbook."
                ) from error
            try:
                with temporary.open("r+b") as stream:
                    os.fsync(stream.fileno())
            except OSError as error:
                if _is_locked_windows_error(error):
                    raise _locked_error() from error
                raise TrackerError(
                    "Could not flush the temporary applications workbook."
                ) from error
            try:
                with temporary.open("rb") as stream:
                    validation_workbook = load_workbook(
                        stream,
                        data_only=False,
                        read_only=False,
                    )
            except PermissionError as error:
                raise _locked_error() from error
            except (
                InvalidFileException,
                KeyError,
                OSError,
                ParseError,
                ValueError,
                zipfile.BadZipFile,
            ) as error:
                if isinstance(error, OSError) and _is_locked_windows_error(error):
                    raise _locked_error() from error
                raise TrackerError(
                    "The temporary workbook could not be validated."
                ) from error
            try:
                self._validated_rows(validation_workbook)
            except TrackerError as error:
                raise TrackerError(
                    "The temporary workbook failed validation."
                ) from error
            finally:
                validation_workbook.close()
            try:
                os.replace(temporary, target)
            except PermissionError as error:
                raise _locked_error() from error
            except OSError as error:
                if _is_locked_windows_error(error):
                    raise _locked_error() from error
                raise TrackerError(
                    "Could not publish the applications workbook."
                ) from error
        finally:
            try:
                temporary.unlink(missing_ok=True)
            except OSError:
                LOGGER.warning("Temporary tracker cleanup failed")


def _validate_application_id(value: object) -> str:
    if (
        not isinstance(value, str)
        or not value
        or len(value) > _MAX_APPLICATION_ID_CHARS
        or value != value.strip()
        or any(unicodedata.category(character) == "Cc" for character in value)
    ):
        raise TrackerValidationError(
            "The application ID must be 1–128 control-free characters with no "
            "surrounding whitespace."
        )
    return value


def _validate_text(value: object, label: str, *, required: bool) -> str:
    if not isinstance(value, str):
        raise TrackerValidationError(f"The {label} must be text.")
    if required and not value.strip():
        raise TrackerValidationError(f"The {label} must not be blank.")
    if len(value) > _MAX_CELL_TEXT_CHARS:
        raise TrackerValidationError(
            f"The {label} exceeds Excel's safe cell-text limit."
        )
    if any(
        ord(character) < 32 and character not in ("\t", "\n", "\r")
        for character in value
    ):
        raise TrackerValidationError(
            f"The {label} contains an unsupported control character."
        )
    return value


def _validate_optional_text(value: object, label: str) -> str | None:
    if value is None or value == "":
        return None
    return _validate_text(value, label, required=False)


def _validate_language(value: object) -> Literal["de", "en"]:
    if value == "de":
        return "de"
    if value == "en":
        return "en"
    raise TrackerValidationError(
        "The application language must be exactly 'de' or 'en'."
    )


def _validate_hash(value: object, label: str) -> str:
    if not isinstance(value, str) or _SHA256_PATTERN.fullmatch(value) is None:
        raise TrackerValidationError(
            f"The {label} must be a lowercase SHA-256 hash."
        )
    return value


def _validate_optional_url(value: object) -> str | None:
    if value is None or value == "":
        return None
    validated = _validate_text(value, "job URL", required=True)
    if validate_http_url(validated) != validated:
        raise TrackerValidationError(
            "The job URL must be a clean absolute HTTP(S) URL or blank."
        )
    return validated


def _validate_requested_date(value: object) -> date | None:
    if value is None:
        return None
    if type(value) is not date:
        raise TrackerValidationError(
            "The applied date must be a calendar date or blank."
        )
    return value


def _validate_stored_date(value: object) -> date | None:
    if value is None:
        return None
    if type(value) is not date:
        raise TrackerValidationError(
            "The stored applied date must be a native Excel calendar date."
        )
    return value


def _parse_utc_timestamp(value: object, label: str) -> datetime:
    if not isinstance(value, str):
        raise TrackerValidationError(
            f"The {label} must be UTC ISO-8601 text."
        )
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as error:
        raise TrackerValidationError(
            f"The {label} must be valid UTC ISO-8601 text."
        ) from error
    if (
        parsed.tzinfo is None
        or parsed.utcoffset() is None
        or parsed.utcoffset() != timedelta(0)
    ):
        raise TrackerValidationError(
            f"The {label} must include the UTC offset."
        )
    return parsed.astimezone(UTC)


def _write_text_cell(cell: Cell, value: object) -> None:
    if not isinstance(value, str):
        raise TrackerValidationError("Tracker text cells require string values.")
    cell.value = value
    cell.data_type = "s"
    cell.hyperlink = None


def _is_locked_windows_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(
        error,
        "winerror",
        None,
    ) in _LOCKED_WINERRORS


def _locked_error() -> TrackerWorkbookLockedError:
    return TrackerWorkbookLockedError(
        "Close applications.xlsx in Excel, then retry."
    )
