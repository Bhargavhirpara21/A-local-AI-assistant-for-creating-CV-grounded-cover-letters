"""Tests for the typed, atomic Excel application tracker."""

from __future__ import annotations

import dataclasses
import logging
import os
import tempfile
import unittest
import zipfile
from copy import copy
from io import StringIO
from dataclasses import dataclass, replace
from datetime import UTC, date, datetime, timedelta, timezone
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from config import Settings, build_settings
from core.tracker import (
    APPLICATION_COLUMNS,
    ApplicationRecord,
    ApplicationTracker,
    TrackerApplicationNotFoundError,
    TrackerError,
    TrackerSchemaError,
    TrackerValidationError,
    TrackerWorkbookLockedError,
)


@dataclass(frozen=True, slots=True)
class _LetterOutputStub:
    """Supply generated metadata consumed by the tracker boundary."""

    company: str
    role: str
    reference_number: str | None
    location: str | None
    job_url: str | None
    language: str
    fit_assessment: str
    contact_person: str | None
    source_hash: str
    input_hash: str
    cv_version_id: str
    cv_reference_hash: str
    used_previous_cv: bool


@dataclass(slots=True)
class _Clock:
    """Return deterministic timezone-aware values in sequence."""

    values: list[datetime]

    def __call__(self) -> datetime:
        """Return and consume the next configured time."""

        return self.values.pop(0)


def _workbook_source_path(source: object) -> Path:
    """Return the filesystem path from a path or managed binary stream."""

    candidate = getattr(source, "name", source)
    if not isinstance(candidate, (str, os.PathLike)):
        raise AssertionError("Workbook source has no filesystem path.")
    return Path(candidate)


class ApplicationTrackerTests(unittest.TestCase):
    """Verify schema, upserts, manual edits, and failure-safe persistence."""

    _temporary_directory: tempfile.TemporaryDirectory[str]
    root: Path
    settings: Settings
    now: datetime
    later: datetime
    output: _LetterOutputStub

    def setUp(self) -> None:
        """Create an isolated tracker and fictional generated metadata."""

        self._temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self._temporary_directory.name).resolve()
        self.settings = build_settings(self.root)
        self.now = datetime(2026, 7, 27, 8, 15, tzinfo=UTC)
        self.later = datetime(2026, 7, 27, 9, 45, tzinfo=UTC)
        self.output = _LetterOutputStub(
            company="Münchner Prüfwerke GmbH",
            role="Automatisierungsingenieur",
            reference_number="DE-Ä42",
            location="München",
            job_url="https://example.test/jobs/%C3%A4-42",
            language="de",
            fit_assessment="Strong match",
            contact_person="Dr. Jörg Weiß",
            source_hash="a" * 64,
            input_hash="b" * 64,
            cv_version_id="cv-fictional-v1",
            cv_reference_hash="c" * 64,
            used_previous_cv=False,
        )

    def tearDown(self) -> None:
        """Release the isolated tracker directory."""

        self._temporary_directory.cleanup()

    def _tracker(self, *times: datetime) -> ApplicationTracker:
        """Build a tracker with a deterministic clock."""

        values = list(times or (self.now,))
        return ApplicationTracker(self.settings, clock=_Clock(values))

    def _letter_path(self, filename: str = "letter.md") -> Path:
        """Create an archive and return its project-relative path."""

        path = self.settings.letters_dir / filename
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("Fictional archived letter.\n", encoding="utf-8")
        return path.relative_to(self.root)

    def _write_workbook(
        self,
        headers: tuple[object, ...],
        rows: tuple[tuple[object, ...], ...] = (),
    ) -> None:
        """Write a direct workbook fixture without using tracker code."""

        self.settings.applications_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Applications"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(self.settings.applications_path)
        workbook.close()

    def _row_values(
        self,
        application_id: str = "app-001",
        *,
        created_at: object | None = None,
        applied_date: object = None,
        status: object = "Draft",
        used_previous_cv: object = False,
    ) -> tuple[object, ...]:
        """Return one structurally complete direct-workbook row fixture."""

        timestamp = created_at if created_at is not None else self.now.isoformat()
        self._letter_path()
        return (
            application_id,
            timestamp,
            timestamp,
            applied_date,
            self.output.company,
            self.output.role,
            self.output.reference_number,
            self.output.location,
            self.output.job_url,
            self.output.language,
            self.output.fit_assessment,
            status,
            self.output.contact_person,
            "letters/letter.md",
            self.output.source_hash,
            self.output.input_hash,
            self.output.cv_version_id,
            self.output.cv_reference_hash,
            used_previous_cv,
            "",
        )

    def test_creates_exact_schema_and_native_typed_immutable_row(self) -> None:
        """First generation should atomically create all 20 typed columns."""

        tracker = self._tracker(self.now)
        letter_path = self._letter_path()

        record = tracker.upsert_generated(
            "app-001",
            self.output,
            letter_path,
        )

        self.assertEqual(
            APPLICATION_COLUMNS,
            (
                "application_id",
                "created_at",
                "updated_at",
                "applied_date",
                "company",
                "role",
                "reference_number",
                "location",
                "job_url",
                "language",
                "fit_assessment",
                "status",
                "contact_person",
                "letter_path",
                "source_hash",
                "input_hash",
                "cv_version_id",
                "cv_reference_hash",
                "used_previous_cv",
                "notes",
            ),
        )
        self.assertIsInstance(record, ApplicationRecord)
        self.assertEqual(record.created_at, self.now)
        self.assertEqual(record.updated_at, self.now)
        self.assertIs(record.created_at.tzinfo, UTC)
        self.assertIs(record.updated_at.tzinfo, UTC)
        self.assertIsNone(record.applied_date)
        self.assertEqual(record.status, "Draft")
        self.assertFalse(record.used_previous_cv)
        self.assertEqual(record.letter_path, "letters/letter.md")
        with self.assertRaises(dataclasses.FrozenInstanceError):
            record.status = "Applied"  # type: ignore[misc]

        workbook = load_workbook(self.settings.applications_path)
        sheet = workbook["Applications"]
        self.assertEqual(
            tuple(cell.value for cell in sheet[1]),
            APPLICATION_COLUMNS,
        )
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(2, 2).value, self.now.isoformat())
        self.assertEqual(sheet.cell(2, 3).value, self.now.isoformat())
        self.assertEqual(sheet.cell(2, 2).data_type, "s")
        self.assertEqual(sheet.cell(2, 3).data_type, "s")
        self.assertIsNone(sheet.cell(2, 4).value)
        self.assertIs(sheet.cell(2, 19).value, False)
        self.assertTrue(
            all(cell.hyperlink is None for cell in sheet[2]),
        )
        workbook.close()

    def test_same_application_id_is_updated_without_duplicate(self) -> None:
        """A refinement should update its existing row and preserve creation."""

        tracker = self._tracker(self.now, self.later, self.later)
        first_path = self._letter_path("initial.md")
        refined_path = self._letter_path("refined.md")
        first = tracker.upsert_generated("app-001", self.output, first_path)
        refined_output = replace(
            self.output,
            company="Münchner Prüfwerke AG",
            input_hash="d" * 64,
            cv_version_id="cv-fictional-v2",
            cv_reference_hash="e" * 64,
            used_previous_cv=True,
        )

        second = tracker.upsert_generated(
            "app-001",
            refined_output,
            refined_path,
        )

        self.assertEqual(first.created_at, second.created_at)
        self.assertEqual(second.updated_at, self.later)
        self.assertEqual(second.company, "Münchner Prüfwerke AG")
        self.assertEqual(second.letter_path, "letters/refined.md")
        self.assertEqual(second.input_hash, "d" * 64)
        self.assertEqual(second.cv_version_id, "cv-fictional-v2")
        self.assertTrue(second.used_previous_cv)
        self.assertEqual(len(tracker.list_applications()), 1)

        workbook = load_workbook(self.settings.applications_path)
        self.assertEqual(workbook["Applications"].max_row, 2)
        workbook.close()

    def test_application_ids_are_strict_bounded_and_case_sensitive(self) -> None:
        """IDs must be exact safe identifiers; casing identifies distinct rows."""

        tracker = self._tracker(self.now, self.later, self.later)
        maximum_id = "x" * 128
        first = tracker.upsert_generated(
            "Case-ID",
            self.output,
            self._letter_path("case-upper.md"),
        )
        second = tracker.upsert_generated(
            "case-ID",
            replace(self.output, input_hash="d" * 64),
            self._letter_path("case-lower.md"),
        )
        maximum = tracker.upsert_generated(
            maximum_id,
            replace(self.output, input_hash="e" * 64),
            self._letter_path("maximum.md"),
        )

        self.assertEqual(first.application_id, "Case-ID")
        self.assertEqual(second.application_id, "case-ID")
        self.assertEqual(maximum.application_id, maximum_id)
        self.assertEqual(len(tracker.list_applications()), 3)

        invalid_ids = (
            "",
            " app-001",
            "app-001 ",
            "line\nbreak",
            "tab\tinside",
            "x" * 129,
        )
        before = self.settings.applications_path.read_bytes()
        for invalid_id in invalid_ids:
            with self.subTest(application_id=repr(invalid_id)):
                with self.assertRaisesRegex(
                    TrackerValidationError,
                    "application ID",
                ):
                    tracker.upsert_generated(
                        invalid_id,
                        replace(self.output, input_hash="f" * 64),
                        self._letter_path("invalid.md"),
                    )
        self.assertEqual(self.settings.applications_path.read_bytes(), before)

    def test_clock_is_converted_to_utc_and_naive_time_fails_before_write(
        self,
    ) -> None:
        """Tracker timestamps must be aware UTC and never guessed from naive time."""

        berlin_time = datetime(
            2026,
            7,
            27,
            10,
            15,
            tzinfo=timezone(timedelta(hours=2)),
        )
        record = self._tracker(berlin_time).upsert_generated(
            "app-aware",
            self.output,
            self._letter_path("aware.md"),
        )
        self.assertEqual(
            record.created_at,
            datetime(2026, 7, 27, 8, 15, tzinfo=UTC),
        )
        self.settings.applications_path.unlink()

        with self.assertRaisesRegex(TrackerValidationError, "timezone-aware"):
            self._tracker(berlin_time.replace(tzinfo=None)).upsert_generated(
                "app-naive",
                self.output,
                self._letter_path("naive.md"),
            )
        self.assertFalse(self.settings.applications_path.exists())

    def test_automated_upsert_preserves_manual_date_status_and_notes(self) -> None:
        """Generation/refinement must never overwrite user-owned fields."""

        tracker = self._tracker(self.now, self.later, self.later)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path("initial.md"),
        )
        edited = tracker.update_manual_fields(
            "app-001",
            applied_date=date(2026, 7, 28),
            status="Applied",
            notes="Über Portal gesendet — Rückfrage offen.",
        )

        updated = tracker.upsert_generated(
            "app-001",
            replace(self.output, input_hash="f" * 64),
            self._letter_path("refined.md"),
        )

        self.assertEqual(updated.applied_date, edited.applied_date)
        self.assertEqual(updated.status, "Applied")
        self.assertEqual(updated.notes, "Über Portal gesendet — Rückfrage offen.")
        self.assertEqual(updated.input_hash, "f" * 64)

    def test_manual_edit_is_explicit_validated_and_does_not_duplicate(self) -> None:
        """One explicit save should update only editable fields on the row."""

        tracker = self._tracker(self.now, self.later)
        original = tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path(),
        )

        updated = tracker.update_manual_fields(
            "app-001",
            applied_date=date(2026, 7, 28),
            status="Interview",
            notes="Gespräch mit Frau Weiß am 31.07.",
        )

        self.assertEqual(updated.application_id, original.application_id)
        self.assertEqual(updated.created_at, original.created_at)
        self.assertEqual(updated.updated_at, self.later)
        self.assertEqual(updated.company, original.company)
        self.assertEqual(updated.applied_date, date(2026, 7, 28))
        self.assertEqual(updated.status, "Interview")
        self.assertEqual(updated.notes, "Gespräch mit Frau Weiß am 31.07.")
        self.assertEqual(len(tracker.list_applications()), 1)

        workbook = load_workbook(self.settings.applications_path)
        applied_cell = workbook["Applications"].cell(2, 4)
        self.assertIsInstance(applied_cell.value, date)
        self.assertNotIsInstance(applied_cell.value, datetime)
        self.assertEqual(applied_cell.value, date(2026, 7, 28))
        workbook.close()

    def test_manual_edit_rejects_unknown_status_invalid_date_and_control_text(
        self,
    ) -> None:
        """Invalid UI edits should be rejected before the workbook is changed."""

        tracker = self._tracker(self.now)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path(),
        )
        original_bytes = self.settings.applications_path.read_bytes()

        with self.assertRaisesRegex(
            TrackerValidationError,
            "valid application status",
        ):
            tracker.update_manual_fields(
                "app-001",
                applied_date=None,
                status="Maybe",
                notes="No change",
            )
        with self.assertRaisesRegex(TrackerValidationError, "calendar date"):
            tracker.update_manual_fields(
                "app-001",
                applied_date="2026-07-28",  # type: ignore[arg-type]
                status="Draft",
                notes="No change",
            )
        with self.assertRaisesRegex(TrackerValidationError, "calendar date"):
            tracker.update_manual_fields(
                "app-001",
                applied_date=datetime(2026, 7, 28),
                status="Draft",
                notes="No change",
            )
        with self.assertRaisesRegex(TrackerValidationError, "unsupported control"):
            tracker.update_manual_fields(
                "app-001",
                applied_date=None,
                status="Draft",
                notes="Unsafe\x01note",
            )

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )

    def test_manual_edit_rejects_unknown_application(self) -> None:
        """An edit cannot silently create a row without generated provenance."""

        tracker = self._tracker(self.now)

        with self.assertRaisesRegex(
            TrackerApplicationNotFoundError,
            "app-missing",
        ):
            tracker.update_manual_fields(
                "app-missing",
                applied_date=None,
                status="Draft",
                notes="",
            )
        self.assertFalse(self.settings.applications_path.exists())

    def test_existing_invalid_timestamp_date_or_boolean_types_fail_closed(
        self,
    ) -> None:
        """Tracker-native fields must not be silently coerced from other types."""

        invalid_rows = (
            self._row_values(created_at=self.now.replace(tzinfo=None)),
            self._row_values(applied_date="2026-07-28"),
            self._row_values(used_previous_cv=1),
        )
        for index, invalid_row in enumerate(invalid_rows):
            with self.subTest(index=index):
                self._write_workbook(APPLICATION_COLUMNS, (invalid_row,))
                before = self.settings.applications_path.read_bytes()
                with self.assertRaises(TrackerSchemaError):
                    self._tracker(self.later).list_applications()
                self.assertEqual(
                    self.settings.applications_path.read_bytes(),
                    before,
                )

    def test_existing_invalid_language_status_hash_or_url_fails_closed(
        self,
    ) -> None:
        """Invalid generated fields in a workbook must never be normalized."""

        corruptions = (
            (9, "fr"),
            (11, "Maybe"),
            (14, "NOT-A-SHA256"),
            (8, "file:///C:/private.txt"),
        )
        for column_index, corrupt_value in corruptions:
            with self.subTest(column=column_index):
                row = list(self._row_values())
                row[column_index] = corrupt_value
                self._write_workbook(APPLICATION_COLUMNS, (tuple(row),))
                before = self.settings.applications_path.read_bytes()

                with self.assertRaises(TrackerSchemaError):
                    self._tracker(self.later).list_applications()

                self.assertEqual(
                    self.settings.applications_path.read_bytes(),
                    before,
                )

    def test_utf8_and_formula_like_text_round_trip_as_literal_strings(self) -> None:
        """Unicode and formula-like prefixes must remain plain unlinked text."""

        tracker = self._tracker(self.now, self.later)
        formula_like = replace(
            self.output,
            company="=HYPERLINK(\"https://invalid.test\",\"München\")",
            role="+Prüfingenieur — Öl & Gas",
            reference_number="-DE-Ä42",
            location="@München",
            contact_person="  =SUM(1,1) Dr. Jörg Weiß",
        )
        tracker.upsert_generated(
            "app-ümlaut",
            formula_like,
            self._letter_path("größe.md"),
        )
        tracker.update_manual_fields(
            "app-ümlaut",
            applied_date=None,
            status="Draft",
            notes="  =1+1 — bleibt Text; Grüße",
        )

        loaded = tracker.list_applications()[0]
        self.assertEqual(loaded.company, formula_like.company)
        self.assertEqual(loaded.role, "+Prüfingenieur — Öl & Gas")
        self.assertEqual(loaded.notes, "  =1+1 — bleibt Text; Grüße")

        workbook = load_workbook(
            self.settings.applications_path,
            data_only=False,
        )
        sheet = workbook["Applications"]
        text_columns = tuple(
            index
            for index, column in enumerate(APPLICATION_COLUMNS, start=1)
            if column not in ("applied_date", "used_previous_cv")
        )
        self.assertTrue(
            all(sheet.cell(2, index).data_type == "s" for index in text_columns),
        )
        self.assertTrue(
            all(sheet.cell(2, index).hyperlink is None for index in text_columns),
        )
        workbook.close()

    def test_existing_formula_or_hyperlink_data_fails_closed(self) -> None:
        """Formulas and hyperlinks in owned application cells are corruption."""

        for corruption in ("formula", "hyperlink"):
            with self.subTest(corruption=corruption):
                self._write_workbook(
                    APPLICATION_COLUMNS,
                    (self._row_values(),),
                )
                workbook = load_workbook(self.settings.applications_path)
                cell = workbook["Applications"].cell(2, 5)
                if corruption == "formula":
                    cell.value = '=HYPERLINK("https://invalid.test","click")'
                else:
                    cell.hyperlink = "https://invalid.test"
                workbook.save(self.settings.applications_path)
                workbook.close()
                before = self.settings.applications_path.read_bytes()

                with self.assertRaisesRegex(
                    TrackerSchemaError,
                    "formula|hyperlink",
                ):
                    self._tracker(self.later).upsert_generated(
                        "app-001",
                        replace(self.output, input_hash="d" * 64),
                        self._letter_path("refined.md"),
                    )

                self.assertEqual(
                    self.settings.applications_path.read_bytes(),
                    before,
                )

    def test_blank_row_hyperlink_is_not_skipped_as_empty_data(self) -> None:
        """A hidden hyperlink in an otherwise blank row must fail closed."""

        self._write_workbook(APPLICATION_COLUMNS)
        with zipfile.ZipFile(self.settings.applications_path, "r") as source:
            entries = tuple(
                (entry, source.read(entry.filename))
                for entry in source.infolist()
            )
        with zipfile.ZipFile(self.settings.applications_path, "w") as target:
            for entry, content in entries:
                if entry.filename == "xl/worksheets/sheet1.xml":
                    content = content.replace(
                        b"</worksheet>",
                        (
                            b'<hyperlinks><hyperlink ref="E2"/>'
                            b"</hyperlinks></worksheet>"
                        ),
                        1,
                    )
                target.writestr(entry, content)
        fixture = load_workbook(self.settings.applications_path)
        self.assertIsNone(fixture["Applications"].cell(2, 5).value)
        self.assertIsNotNone(fixture["Applications"].cell(2, 5).hyperlink)
        fixture.close()
        original_bytes = self.settings.applications_path.read_bytes()

        with self.assertRaisesRegex(TrackerSchemaError, "hyperlink"):
            self._tracker(self.now).list_applications()

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )

    def test_blank_optional_text_is_canonicalized_to_none(self) -> None:
        """Empty optional metadata should have one stable workbook meaning."""

        tracker = self._tracker(self.now)
        blank_optional = replace(
            self.output,
            reference_number="",
            location="",
            job_url="",
            contact_person="",
        )

        record = tracker.upsert_generated(
            "app-blank-optional",
            blank_optional,
            self._letter_path(),
        )
        loaded = tracker.list_applications()[0]

        self.assertIsNone(record.reference_number)
        self.assertIsNone(record.location)
        self.assertIsNone(record.job_url)
        self.assertIsNone(record.contact_person)
        self.assertEqual(record, loaded)

    def test_oversized_generated_url_is_rejected_before_excel_truncation(
        self,
    ) -> None:
        """Excel must never silently shorten a generated application URL."""

        oversized_url = "https://example.test/" + ("a" * 32_768)
        tracker = self._tracker(self.now)

        with self.assertRaisesRegex(TrackerValidationError, "safe cell-text"):
            tracker.upsert_generated(
                "app-oversized-url",
                replace(self.output, job_url=oversized_url),
                self._letter_path(),
            )

        self.assertFalse(self.settings.applications_path.exists())

    def test_unrelated_worksheets_are_preserved_without_tracker_links(self) -> None:
        """Saving Applications must retain unrelated workbook-owned content."""

        self.settings.applications_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        applications = workbook.active
        applications.title = "Applications"
        applications.append(APPLICATION_COLUMNS)
        notes_sheet = workbook.create_sheet("Private Notes")
        notes_sheet["B3"] = "Do not replace — Grüße"
        bold_font = copy(notes_sheet["B3"].font)
        bold_font.bold = True
        notes_sheet["B3"].font = bold_font
        workbook.save(self.settings.applications_path)
        workbook.close()

        self._tracker(self.now).upsert_generated(
            "app-001",
            self.output,
            self._letter_path(),
        )

        loaded = load_workbook(self.settings.applications_path)
        self.assertEqual(
            loaded["Private Notes"]["B3"].value,
            "Do not replace — Grüße",
        )
        self.assertTrue(loaded["Private Notes"]["B3"].font.bold)
        self.assertTrue(
            all(
                cell.hyperlink is None
                for row in loaded["Applications"].iter_rows()
                for cell in row
            ),
        )
        loaded.close()

    def test_letter_paths_are_existing_markdown_archives_confined_to_letters(
        self,
    ) -> None:
        """Only project-relative existing Markdown archives should be accepted."""

        tracker = self._tracker(self.now)
        internal_path = self._letter_path("nested/letter.md")

        internal = tracker.upsert_generated(
            "app-internal",
            self.output,
            internal_path,
        )
        self.assertEqual(internal.letter_path, "letters/nested/letter.md")
        self.assertEqual(
            tracker.resolve_letter_path(internal.letter_path),
            (self.root / internal_path).resolve(),
        )

        with self.assertRaisesRegex(TrackerValidationError, "project directory"):
            tracker.upsert_generated(
                "app-traversal",
                replace(self.output, input_hash="d" * 64),
                Path("..") / "escaped.md",
            )
        with self.assertRaisesRegex(TrackerValidationError, "relative"):
            tracker.upsert_generated(
                "app-absolute",
                replace(self.output, input_hash="e" * 64),
                (self.root / internal_path).resolve(),
            )
        with self.assertRaisesRegex(TrackerValidationError, "does not exist"):
            tracker.upsert_generated(
                "app-missing",
                replace(self.output, input_hash="f" * 64),
                Path("letters") / "missing.md",
            )
        non_markdown = self.settings.letters_dir / "letter.txt"
        non_markdown.write_text("Not a Markdown archive.\n", encoding="utf-8")
        with self.assertRaisesRegex(TrackerValidationError, "Markdown"):
            tracker.upsert_generated(
                "app-non-markdown",
                replace(self.output, input_hash="1" * 64),
                non_markdown.relative_to(self.root),
            )

    def test_symlink_escape_from_letters_is_rejected(self) -> None:
        """A relative archive symlink must not escape the private archive."""

        tracker = self._tracker(self.now)
        outside_path = self.root / "outside.md"
        outside_path.write_text("Outside archive.\n", encoding="utf-8")
        link_path = self.settings.letters_dir / "escape.md"
        link_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            link_path.symlink_to(outside_path)
        except OSError as error:
            self.skipTest(f"File symlinks are unavailable: {type(error).__name__}")

        with self.assertRaisesRegex(TrackerValidationError, "letters directory"):
            tracker.upsert_generated(
                "app-symlink",
                self.output,
                link_path.relative_to(self.root),
            )

    def test_rejects_wrong_or_duplicate_schema_without_overwriting(self) -> None:
        """Existing workbooks with altered headers must fail closed."""

        bad_headers = list(APPLICATION_COLUMNS)
        bad_headers[1] = "application_id"
        self._write_workbook(tuple(bad_headers))
        original_bytes = self.settings.applications_path.read_bytes()
        tracker = self._tracker(self.now)

        with self.assertRaisesRegex(TrackerSchemaError, "20-column schema"):
            tracker.list_applications()
        with self.assertRaisesRegex(TrackerSchemaError, "20-column schema"):
            tracker.upsert_generated(
                "app-001",
                self.output,
                self._letter_path(),
            )

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )

    def test_rejects_duplicate_application_ids_without_data_loss(self) -> None:
        """A manually corrupted duplicate ID must never be guessed or merged."""

        timestamp = self.now.isoformat()
        self._letter_path()
        row = (
            "app-duplicate",
            timestamp,
            timestamp,
            None,
            self.output.company,
            self.output.role,
            self.output.reference_number,
            self.output.location,
            self.output.job_url,
            self.output.language,
            self.output.fit_assessment,
            "Draft",
            self.output.contact_person,
            "letters/letter.md",
            self.output.source_hash,
            self.output.input_hash,
            self.output.cv_version_id,
            self.output.cv_reference_hash,
            False,
            "",
        )
        self._write_workbook(APPLICATION_COLUMNS, (row, row))
        original_bytes = self.settings.applications_path.read_bytes()
        tracker = self._tracker(self.later)

        with self.assertRaisesRegex(
            TrackerSchemaError,
            "Duplicate application ID",
        ):
            tracker.list_applications()

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )

    def test_rejects_corrupt_workbook_without_replacing_it(self) -> None:
        """Non-XLSX content should produce a repair message and remain intact."""

        self.settings.applications_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )
        corrupt_bytes = b"not an xlsx workbook"
        self.settings.applications_path.write_bytes(corrupt_bytes)
        tracker = self._tracker(self.now)

        with self.assertRaisesRegex(TrackerSchemaError, "invalid or corrupted"):
            tracker.upsert_generated(
                "app-001",
                self.output,
                self._letter_path(),
            )

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            corrupt_bytes,
        )

    def test_rejects_malformed_internal_xml_without_replacing_it(self) -> None:
        """Broken XML inside an XLSX should map to the corruption boundary."""

        self._write_workbook(APPLICATION_COLUMNS)
        with zipfile.ZipFile(self.settings.applications_path, "r") as source:
            entries = tuple(
                (entry, source.read(entry.filename))
                for entry in source.infolist()
            )
        with zipfile.ZipFile(self.settings.applications_path, "w") as target:
            for entry, content in entries:
                if entry.filename == "xl/worksheets/sheet1.xml":
                    content = b"<worksheet><broken>"
                target.writestr(entry, content)
        corrupt_bytes = self.settings.applications_path.read_bytes()

        with self.assertRaisesRegex(TrackerSchemaError, "invalid or corrupted"):
            self._tracker(self.now).list_applications()

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            corrupt_bytes,
        )

    def test_atomic_save_uses_temporary_file_in_workbook_directory(self) -> None:
        """A same-dir .xlsx must be reopened successfully before publication."""

        tracker = self._tracker(self.now)
        real_replace = os.replace
        real_load = load_workbook
        events: list[tuple[str, Path, Path | None]] = []

        def recording_load(
            filename: str | os.PathLike[str],
            *args: object,
            **kwargs: object,
        ) -> object:
            """Record validation reopening and delegate to openpyxl."""

            path = _workbook_source_path(filename)
            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 0)
            events.append(("load", path, None))
            return real_load(filename, *args, **kwargs)

        def recording_replace(source: str | os.PathLike[str], destination: str | os.PathLike[str]) -> None:
            """Record and execute an atomic replacement."""

            source_path = Path(source)
            destination_path = Path(destination)
            events.append(("replace", source_path, destination_path))
            real_replace(source_path, destination_path)

        with (
            mock.patch(
                "core.tracker.load_workbook",
                side_effect=recording_load,
            ),
            mock.patch(
                "core.tracker.os.replace",
                side_effect=recording_replace,
            ),
        ):
            tracker.upsert_generated(
                "app-001",
                self.output,
                self._letter_path(),
            )

        self.assertEqual([event[0] for event in events], ["load", "replace"])
        validation_path = events[0][1]
        source = events[1][1]
        destination = events[1][2]
        self.assertEqual(validation_path, source)
        self.assertEqual(source.suffix, ".xlsx")
        self.assertEqual(source.parent, self.settings.applications_path.parent)
        self.assertEqual(destination, self.settings.applications_path)
        self.assertFalse(source.exists())

    def test_temporary_save_or_validation_failure_never_publishes(self) -> None:
        """Failures before replace should clean temporary files and keep target."""

        tracker = self._tracker(self.now, self.later, self.later)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path("initial.md"),
        )
        original_bytes = self.settings.applications_path.read_bytes()

        with mock.patch(
            "core.tracker.Workbook.save",
            side_effect=OSError("simulated disk failure"),
        ):
            with self.assertRaisesRegex(TrackerError, "save"):
                tracker.upsert_generated(
                    "app-001",
                    replace(self.output, input_hash="d" * 64),
                    self._letter_path("refined.md"),
                )
        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )

        real_load = load_workbook

        def fail_temporary_reopen(
            filename: str | os.PathLike[str],
            *args: object,
            **kwargs: object,
        ) -> object:
            """Load the target but simulate an invalid temporary workbook."""

            if _workbook_source_path(filename) == self.settings.applications_path:
                return real_load(filename, *args, **kwargs)
            raise zipfile.BadZipFile("simulated invalid temporary workbook")

        with mock.patch(
            "core.tracker.load_workbook",
            side_effect=fail_temporary_reopen,
        ):
            with self.assertRaisesRegex(TrackerError, "temporary workbook"):
                tracker.upsert_generated(
                    "app-001",
                    replace(self.output, input_hash="e" * 64),
                    self._letter_path("refined-again.md"),
                )
        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )
        self.assertEqual(
            tuple(
                self.settings.applications_path.parent.glob(
                    ".applications-*.xlsx",
                )
            ),
            (),
        )

    def test_locked_workbook_error_is_actionable_and_preserves_existing_file(
        self,
    ) -> None:
        """A Windows replace lock should retain the previous valid workbook."""

        tracker = self._tracker(self.now, self.later)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path("initial.md"),
        )
        original_bytes = self.settings.applications_path.read_bytes()

        with mock.patch(
            "core.tracker.os.replace",
            side_effect=PermissionError(13, "file is in use"),
        ):
            with self.assertRaisesRegex(
                TrackerWorkbookLockedError,
                "Close applications.xlsx in Excel",
            ):
                tracker.upsert_generated(
                    "app-001",
                    replace(self.output, input_hash="d" * 64),
                    self._letter_path("refined.md"),
                )

        self.assertEqual(
            self.settings.applications_path.read_bytes(),
            original_bytes,
        )
        self.assertEqual(
            tuple(
                self.settings.applications_path.parent.glob(
                    ".applications-*.xlsx",
                )
            ),
            (),
        )

    def test_locked_workbook_read_is_actionable(self) -> None:
        """A Windows read lock should provide the same direct recovery step."""

        tracker = self._tracker(self.now)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path(),
        )

        with mock.patch(
            "core.tracker.load_workbook",
            side_effect=PermissionError(13, "file is in use"),
        ):
            with self.assertRaisesRegex(
                TrackerWorkbookLockedError,
                "Close applications.xlsx in Excel",
            ):
                tracker.list_applications()

    def test_windows_sharing_error_codes_receive_close_excel_guidance(self) -> None:
        """WinError 5, 32, and 33 must all map to one actionable lock error."""

        tracker = self._tracker(self.now)
        tracker.upsert_generated(
            "app-001",
            self.output,
            self._letter_path(),
        )

        for winerror in (5, 32, 33):
            with self.subTest(winerror=winerror):
                failure = OSError("simulated Windows sharing violation")
                failure.winerror = winerror  # type: ignore[attr-defined]
                with mock.patch(
                    "core.tracker.load_workbook",
                    side_effect=failure,
                ):
                    with self.assertRaisesRegex(
                        TrackerWorkbookLockedError,
                        "Close applications.xlsx in Excel",
                    ):
                        tracker.list_applications()

    def test_temporary_flush_and_validation_locks_are_actionable(self) -> None:
        """Temporary-file sharing errors should retain close-Excel guidance."""

        self._tracker(self.now).upsert_generated(
            "app-001",
            self.output,
            self._letter_path("initial.md"),
        )
        original_bytes = self.settings.applications_path.read_bytes()
        real_load = load_workbook

        for stage in ("flush", "validation"):
            with self.subTest(stage=stage):
                failure = OSError("simulated Windows sharing violation")
                failure.winerror = 32  # type: ignore[attr-defined]

                def temporary_load(
                    filename: str | os.PathLike[str],
                    *args: object,
                    **kwargs: object,
                ) -> object:
                    """Load the target but fail reopening its temporary copy."""

                    if (
                        _workbook_source_path(filename)
                        == self.settings.applications_path
                    ):
                        return real_load(filename, *args, **kwargs)
                    raise failure

                patcher = (
                    mock.patch("core.tracker.os.fsync", side_effect=failure)
                    if stage == "flush"
                    else mock.patch(
                        "core.tracker.load_workbook",
                        side_effect=temporary_load,
                    )
                )
                with patcher:
                    with self.assertRaisesRegex(
                        TrackerWorkbookLockedError,
                        "Close applications.xlsx in Excel",
                    ):
                        self._tracker(self.later).upsert_generated(
                            "app-001",
                            replace(self.output, input_hash="d" * 64),
                            self._letter_path(f"{stage}.md"),
                        )

                self.assertEqual(
                    self.settings.applications_path.read_bytes(),
                    original_bytes,
                )
                self.assertEqual(
                    tuple(
                        self.settings.applications_path.parent.glob(
                            ".applications-*.xlsx",
                        )
                    ),
                    (),
                )

    def test_operational_logs_never_include_application_row_content(self) -> None:
        """Even failure logs must not expose identifiers or generated metadata."""

        tracker = self._tracker(self.now, self.later, self.later)
        sensitive_id = "PRIVATE-APPLICATION-ID"
        sensitive_company = "PRIVATE-COMPANY-NAME"
        sensitive_url = "https://private.example.test/jobs/secret"
        sensitive_notes = "PRIVATE INTERVIEW NOTES"
        sensitive_output = replace(
            self.output,
            company=sensitive_company,
            job_url=sensitive_url,
        )
        tracker.upsert_generated(
            sensitive_id,
            sensitive_output,
            self._letter_path(),
        )
        tracker.update_manual_fields(
            sensitive_id,
            applied_date=None,
            status="Draft",
            notes=sensitive_notes,
        )
        stream = StringIO()
        handler = logging.StreamHandler(stream)
        tracker_logger = logging.getLogger("core.tracker")
        tracker_logger.addHandler(handler)
        previous_level = tracker_logger.level
        tracker_logger.setLevel(logging.INFO)
        try:
            with mock.patch(
                "core.tracker.os.replace",
                side_effect=PermissionError(13, "file is in use"),
            ):
                with self.assertRaises(TrackerWorkbookLockedError):
                    tracker.upsert_generated(
                        sensitive_id,
                        replace(sensitive_output, input_hash="d" * 64),
                        self._letter_path("refined.md"),
                    )
        finally:
            tracker_logger.removeHandler(handler)
            tracker_logger.setLevel(previous_level)
            handler.close()

        logged = stream.getvalue()
        self.assertNotIn(sensitive_id, logged)
        self.assertNotIn(sensitive_company, logged)
        self.assertNotIn(sensitive_url, logged)
        self.assertNotIn(sensitive_notes, logged)
        self.assertNotIn("Dr. Jörg Weiß", logged)
        self.assertNotIn(str(self.root), logged)

    def test_rejects_invalid_generated_provenance_before_writing(self) -> None:
        """Required hashes and boolean fallback provenance must remain typed."""

        tracker = self._tracker(self.now)

        with self.assertRaisesRegex(TrackerValidationError, "source hash"):
            tracker.upsert_generated(
                "app-001",
                replace(self.output, source_hash="not-a-hash"),
                self._letter_path(),
            )
        with self.assertRaisesRegex(TrackerValidationError, "boolean"):
            tracker.upsert_generated(
                "app-001",
                replace(
                    self.output,
                    used_previous_cv=1,  # type: ignore[arg-type]
                ),
                self._letter_path(),
            )
        self.assertFalse(self.settings.applications_path.exists())


if __name__ == "__main__":
    unittest.main()
